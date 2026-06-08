import io
import math
import traceback
import warnings
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageOps

from Altomizer.alt_management import (
    build_alt_preview_entry as build_preview_from_alt_management,
    is_real_visual_inventory_item,
    normalize_alt_text,
)
from Altomizer.converter import docx_to_pdf
from Altomizer.style_inspector import extract_generic_drawing_object_from_node, extract_mathtype_object_from_node

DOCX_XML_NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "wp": "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "v": "urn:schemas-microsoft-com:vml",
    "o": "urn:schemas-microsoft-com:office:office",
    "pic": "http://schemas.openxmlformats.org/drawingml/2006/picture",
    "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
}

for prefix, uri in DOCX_XML_NS.items():
    ET.register_namespace(prefix, uri)

def safe_download_stem(filename: str) -> str:
    stem = Path(filename).stem
    cleaned = "".join(char if char.isalnum() or char in {"-", "_"} else "_" for char in stem)
    cleaned = cleaned.strip("._")
    return cleaned or "alt_inventory"


def media_extension(media_target: str | None) -> str:
    if not media_target:
        return "png"
    name = Path(media_target).name
    suffix = Path(name).suffix.lower().lstrip(".")
    return suffix or "png"


def media_type_from_extension(ext: str | None) -> str:
    lookup = {
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "gif": "image/gif",
        "bmp": "image/bmp",
        "webp": "image/webp",
    }
    normalized = (ext or "png").lower().lstrip(".")
    return lookup.get(normalized, "image/png")


def build_alt_preview_entry(session: dict, row: dict) -> dict | None:
    preview_images = session.get("preview_images")
    row_id = row.get("id")
    if isinstance(preview_images, dict) and isinstance(row_id, int):
        cached = preview_images.get(row_id)
        if isinstance(cached, dict):
            return cached

    media_source_path = session.get("media_source_path")
    source_path = media_source_path if isinstance(media_source_path, Path) else session.get("source_path")
    pdf_path = session.get("pdf_path")
    return build_preview_from_alt_management(
        row,
        source_path if isinstance(source_path, Path) else None,
        pdf_path if isinstance(pdf_path, Path) else None,
    )


def get_excel_anchor_position(image: object) -> tuple[int, int]:
    try:
        return image.anchor._from.row, image.anchor._from.col
    except Exception:
        return 999999, 999999


def extract_excel_images(workbook_bytes: bytes) -> list[dict]:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
        )
        workbook = load_workbook(io.BytesIO(workbook_bytes))
    extracted = []

    try:
        for sheet in workbook.worksheets:
            for index, xl_image in enumerate(getattr(sheet, "_images", []), start=1):
                try:
                    data = xl_image._data()
                    with Image.open(io.BytesIO(data)) as source_image:
                        pil_image = source_image.convert("RGB")

                    row, col = get_excel_anchor_position(xl_image)
                    extracted.append(
                        {
                            "sheet": sheet.title,
                            "row": row,
                            "col": col,
                            "index": index,
                            "image": pil_image,
                        }
                    )
                except Exception:
                    traceback.print_exc()
    finally:
        workbook.close()

    extracted.sort(key=lambda item: (item["sheet"], item["row"], item["col"], item["index"]))
    return extracted


def make_grids(
    images: list[dict],
    *,
    max_per_grid: int = 24,
    cols: int = 4,
    cell_w: int = 320,
    cell_h: int = 220,
    margin: int = 24,
    bg: str = "white",
) -> list[tuple[str, bytes]]:
    if not images:
        return []

    rows = math.ceil(max_per_grid / cols)
    grid_files = []

    for grid_no in range(math.ceil(len(images) / max_per_grid)):
        batch = images[grid_no * max_per_grid : (grid_no + 1) * max_per_grid]
        canvas_w = cols * cell_w + (cols + 1) * margin
        canvas_h = rows * cell_h + (rows + 1) * margin

        canvas = Image.new("RGB", (canvas_w, canvas_h), bg)
        draw = ImageDraw.Draw(canvas)

        for index, item in enumerate(batch):
            thumb = ImageOps.contain(item["image"], (cell_w - 20, cell_h - 35))
            x = margin + (index % cols) * (cell_w + margin)
            y = margin + (index // cols) * (cell_h + margin)

            tile = Image.new("RGB", (cell_w, cell_h), (245, 245, 245))
            paste_x = (cell_w - thumb.width) // 2
            paste_y = (cell_h - thumb.height) // 2 + 10
            tile.paste(thumb, (paste_x, paste_y))

            canvas.paste(tile, (x, y))
            draw.rectangle([x, y, x + cell_w, y + cell_h], outline="black", width=1)

            label = f"{grid_no * max_per_grid + index + 1} | {item['sheet']} R{item['row'] + 1}C{item['col'] + 1}"
            draw.text((x + 8, y + 6), label, fill="black")

        output = io.BytesIO()
        canvas.save(output, format="PNG")
        grid_files.append((f"grid_{grid_no + 1}.png", output.getvalue()))

    return grid_files


def clear_alt_text_from_docx(docx_path: Path) -> tuple[bytes, int]:
    updated_parts: dict[str, bytes] = {}
    cleared_count = 0

    with zipfile.ZipFile(docx_path, "r") as source_archive:
        members = source_archive.infolist()

        for member in members:
            if not member.filename.startswith("word/") or not member.filename.endswith(".xml"):
                continue

            xml_bytes = source_archive.read(member.filename)
            try:
                root = ET.fromstring(xml_bytes)
            except ET.ParseError:
                continue

            part_cleared = 0

            for doc_pr in root.findall(".//wp:docPr", DOCX_XML_NS):
                for attr in ("descr", "title"):
                    if doc_pr.attrib.pop(attr, None) is not None:
                        part_cleared += 1

            for shape in root.findall(".//v:shape", DOCX_XML_NS):
                for attr in ("alt", "title"):
                    if shape.attrib.pop(attr, None) is not None:
                        part_cleared += 1

            for image_data in root.findall(".//v:imagedata", DOCX_XML_NS):
                for attr in ("title", f"{{{DOCX_XML_NS['o']}}}title"):
                    if image_data.attrib.pop(attr, None) is not None:
                        part_cleared += 1

            for ole_object in root.findall(".//o:OLEObject", DOCX_XML_NS):
                for attr in ("alt", "title", "descr", f"{{{DOCX_XML_NS['o']}}}title"):
                    if ole_object.attrib.pop(attr, None) is not None:
                        part_cleared += 1

            for nv_props in root.findall(".//pic:cNvPr", DOCX_XML_NS):
                for attr in ("descr", "title"):
                    if nv_props.attrib.pop(attr, None) is not None:
                        part_cleared += 1

            if part_cleared:
                updated_parts[member.filename] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                cleared_count += part_cleared

        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target_archive:
            for member in members:
                if member.filename in updated_parts:
                    target_archive.writestr(member, updated_parts[member.filename])
                else:
                    target_archive.writestr(member, source_archive.read(member.filename))

    return output.getvalue(), cleared_count


def parse_alt_injection_workbook(
    workbook_bytes: bytes,
    session_rows: list[dict],
    expected_source_filename: str | None = None,
) -> list[str]:
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    try:
        if not workbook.worksheets:
            raise ValueError("The uploaded ALT workbook does not contain any worksheets.")

        sheet = workbook.worksheets[0]
        header_map: dict[str, int] = {}
        for col_index in range(1, sheet.max_column + 1):
            raw_value = sheet.cell(row=1, column=col_index).value
            if raw_value is None:
                continue
            normalized = " ".join(str(raw_value).strip().lower().split())
            if normalized:
                header_map[normalized] = col_index

        alt_col = header_map.get("alt text") or header_map.get("generated alt text")
        if alt_col is None:
            raise ValueError("The uploaded workbook must include an 'ALT Text' column.")
        item_id_col = header_map.get("item id")
        file_name_col = header_map.get("file name")
        expected_rows = len(session_rows)

        if sheet.max_row < expected_rows + 1:
            raise ValueError(f"The uploaded workbook has {max(0, sheet.max_row - 1)} ALT row(s), but {expected_rows} were expected.")

        alt_by_item_id: dict[int, str] = {}
        alt_texts_in_order: list[str] = []
        workbook_filenames: set[str] = set()
        for row_index in range(2, expected_rows + 2):
            value = sheet.cell(row=row_index, column=alt_col).value
            normalized_alt = "" if value is None else " ".join(str(value).replace("\n", " ").split())
            alt_texts_in_order.append(normalized_alt)
            if item_id_col is not None:
                raw_item_id = sheet.cell(row=row_index, column=item_id_col).value
                try:
                    item_id = int(str(raw_item_id).strip())
                except (TypeError, ValueError):
                    item_id = None
                if item_id is not None:
                    alt_by_item_id[item_id] = normalized_alt
            if file_name_col is not None:
                file_value = sheet.cell(row=row_index, column=file_name_col).value
                normalized_file_value = "" if file_value is None else str(file_value).strip()
                if normalized_file_value:
                    workbook_filenames.add(normalized_file_value)

        normalized_expected = (expected_source_filename or "").strip().lower()
        normalized_workbook_filenames = {name.strip().lower() for name in workbook_filenames if name.strip()}
        if normalized_expected and normalized_workbook_filenames and normalized_workbook_filenames != {normalized_expected}:
            raise ValueError(
                f"The uploaded workbook references {', '.join(sorted(workbook_filenames))}, but the current DOCX is {expected_source_filename}."
            )

        if item_id_col is not None and len(alt_by_item_id) >= expected_rows:
            return [alt_by_item_id.get(int(row.get("id", index)), "") for index, row in enumerate(session_rows)]

        return alt_texts_in_order
    finally:
        workbook.close()


def normalize_visual_target(value: object) -> str:
    if not isinstance(value, str):
        return ""
    return value.strip().replace("\\", "/").lstrip("/").lower()


def row_visual_targets(row: dict) -> set[str]:
    values = {
        normalize_visual_target(row.get("media_target")),
        normalize_visual_target(row.get("ole_target")),
    }
    return {value for value in values if value}


def candidate_visual_targets(candidate: dict) -> set[str]:
    metadata = candidate.get("metadata") or {}
    values = {
        normalize_visual_target(metadata.get("target")),
        normalize_visual_target(metadata.get("image_target")),
        normalize_visual_target(metadata.get("ole_target")),
    }
    return {value for value in values if value}


def match_session_rows_to_injection_targets(session_rows: list[dict], targets: list[dict]) -> list[dict | None]:
    remaining = list(targets)
    matched_targets: list[dict | None] = []

    def pop_match(predicate) -> dict | None:
        for index, target in enumerate(remaining):
            if predicate(target):
                return remaining.pop(index)
        return None

    for row_index, row in enumerate(session_rows, start=1):
        expected_role = str(row.get("role", "") or "").strip().lower()
        expected_source_part = str(row.get("source_part", "") or "").strip().lower()
        expected_targets = row_visual_targets(row)

        target = None
        if expected_targets:
            target = pop_match(
                lambda candidate: candidate.get("role") == expected_role
                and candidate.get("source_part") == expected_source_part
                and bool(candidate_visual_targets(candidate) & expected_targets)
            )
        if target is None and expected_targets:
            target = pop_match(
                lambda candidate: candidate.get("role") == expected_role
                and bool(candidate_visual_targets(candidate) & expected_targets)
            )
        if target is None and expected_targets and expected_source_part:
            target = pop_match(
                lambda candidate: candidate.get("source_part") == expected_source_part
                and bool(candidate_visual_targets(candidate) & expected_targets)
            )
        if target is None and expected_targets:
            target = pop_match(
                lambda candidate: bool(candidate_visual_targets(candidate) & expected_targets)
            )
        if target is None and expected_source_part and expected_role:
            target = pop_match(
                lambda candidate: candidate.get("role") == expected_role
                and candidate.get("source_part") == expected_source_part
            )
        if target is None and expected_role:
            target = pop_match(lambda candidate: candidate.get("role") == expected_role)

        if target is None:
            # Some DOCX equations are discoverable during ALT collection but are
            # native Office Math objects without an ALT-bearing visual node we can
            # write back to. Skip those instead of blocking the whole DOCX export.
            if expected_role == "equation":
                matched_targets.append(None)
                continue
            raise ValueError(
                f"ALT row {row_index} could not be matched back to an injectable {expected_role or 'visual'} in the current DOCX."
            )

        matched_targets.append(target)

    return matched_targets


def resolve_docx_part_target(base_part: str, target: str) -> str:
    if not target:
        return ""
    if target.startswith("/"):
        return target.lstrip("/")
    return str((Path(base_part).parent / target).as_posix())


def build_docx_package_assets(source_archive: zipfile.ZipFile) -> tuple[set[str], set[str]]:
    names = set(source_archive.namelist())
    media_targets = {name for name in names if name.lower().startswith("word/media/")}
    equation_targets = set()

    for name in names:
        lower_name = name.lower()
        if not lower_name.startswith("word/embeddings/"):
            continue
        try:
            sample = source_archive.read(name)[:8192]
        except KeyError:
            continue
        decoded = sample.decode("latin1", errors="ignore").lower()
        if any(marker in decoded for marker in ("equation.dsmt", "mathtype", "microsoft equation 3.0", "equation editor")):
            equation_targets.add(name)

    return media_targets, equation_targets


def build_part_package_info(
    source_archive: zipfile.ZipFile,
    part_name: str,
    media_targets: set[str],
    equation_targets: set[str],
) -> dict:
    relationships = {}
    rels_name = f"{Path(part_name).parent.as_posix()}/_rels/{Path(part_name).name}.rels"

    if rels_name in source_archive.namelist():
        rels_root = ET.fromstring(source_archive.read(rels_name))
        for rel in rels_root.findall("pr:Relationship", {"pr": "http://schemas.openxmlformats.org/package/2006/relationships"}):
            rel_id = rel.attrib.get("Id")
            target = resolve_docx_part_target(part_name, rel.attrib.get("Target", ""))
            rel_type = rel.attrib.get("Type", "")
            if rel_id and target:
                relationships[rel_id] = {"target": target, "type": rel_type}

    return {
        "relationships": relationships,
        "equation_targets": equation_targets,
        "media_targets": media_targets,
    }


def collect_docx_visual_targets(body: ET.Element, package_info: dict | None) -> list[dict]:
    word_paragraph_tag = f"{{{DOCX_XML_NS['w']}}}p"
    word_table_tag = f"{{{DOCX_XML_NS['w']}}}tbl"
    targets: list[dict] = []

    def append_paragraph_targets(paragraph: ET.Element) -> None:
        for node in paragraph.findall(".//w:object", DOCX_XML_NS):
            math_object = extract_mathtype_object_from_node(node, package_info)
            if not math_object:
                continue
            entry = {
                "role": "equation",
                "text": math_object.get("alt_text", ""),
                "metadata": {
                    "prog_id": math_object.get("prog_id"),
                    "image_rid": math_object.get("image_rid"),
                    "ole_rid": math_object.get("ole_rid"),
                    "image_target": math_object.get("image_target"),
                    "ole_target": math_object.get("ole_target"),
                    "width_pt": math_object.get("width_pt"),
                    "height_pt": math_object.get("height_pt"),
                },
                "has_alt_text": math_object.get("has_alt_text"),
            }
            if not is_real_visual_inventory_item(entry):
                continue
            targets.append({"node": node, "role": "equation", "kind": "object", "metadata": entry["metadata"]})

        for node in paragraph.findall(".//w:drawing", DOCX_XML_NS):
            drawing_object = extract_generic_drawing_object_from_node(node, package_info)
            if not drawing_object:
                continue
            role = "equation" if drawing_object.get("is_equation") else "image"
            entry = {
                "role": role,
                "text": drawing_object.get("alt_text", ""),
                "metadata": drawing_object.get("metadata") or {},
                "has_alt_text": drawing_object.get("has_alt_text"),
            }
            if not is_real_visual_inventory_item(entry):
                continue
            targets.append({"node": node, "role": role, "kind": "drawing", "metadata": entry["metadata"]})

        for node in paragraph.findall(".//w:pict", DOCX_XML_NS):
            drawing_object = extract_generic_drawing_object_from_node(node, package_info)
            if not drawing_object:
                continue
            role = "equation" if drawing_object.get("is_equation") else "image"
            entry = {
                "role": role,
                "text": drawing_object.get("alt_text", ""),
                "metadata": drawing_object.get("metadata") or {},
                "has_alt_text": drawing_object.get("has_alt_text"),
            }
            if not is_real_visual_inventory_item(entry):
                continue
            targets.append({"node": node, "role": role, "kind": "pict", "metadata": entry["metadata"]})

    if body.tag == word_paragraph_tag:
        append_paragraph_targets(body)
        return targets

    if body.tag == word_table_tag:
        for paragraph in body.findall(".//w:p", DOCX_XML_NS):
            append_paragraph_targets(paragraph)
        return targets

    for child in body:
        if child.tag == word_paragraph_tag:
            append_paragraph_targets(child)
        elif child.tag == word_table_tag:
            for paragraph in child.findall(".//w:p", DOCX_XML_NS):
                append_paragraph_targets(paragraph)

    return targets


def collect_docx_injection_targets(source_archive: zipfile.ZipFile) -> tuple[dict[str, ET.Element], list[dict]]:
    media_targets, equation_targets = build_docx_package_assets(source_archive)
    part_roots: dict[str, ET.Element] = {}
    targets: list[dict] = []
    names = set(source_archive.namelist())

    part_plan = [
        ("header", sorted(name for name in names if name.startswith("word/header") and name.endswith(".xml"))),
        ("body", ["word/document.xml"] if "word/document.xml" in names else []),
        ("footer", sorted(name for name in names if name.startswith("word/footer") and name.endswith(".xml"))),
    ]

    for source_part, part_names in part_plan:
        for part_name in part_names:
            root = ET.fromstring(source_archive.read(part_name))
            part_roots[part_name] = root
            package_info = build_part_package_info(source_archive, part_name, media_targets, equation_targets)

            if source_part == "body":
                body = root.find(".//w:body", DOCX_XML_NS)
                if body is None:
                    continue
                part_targets = collect_docx_visual_targets(body, package_info)
            else:
                part_targets = []
                for paragraph in root.findall(".//w:p", DOCX_XML_NS):
                    part_targets.extend(collect_docx_visual_targets(paragraph, package_info))

            for item in part_targets:
                item["source_part"] = source_part
                item["part_name"] = part_name
            targets.extend(part_targets)

    return part_roots, targets


def apply_alt_text_to_visual_node(node: ET.Element, alt_text: str) -> None:
    normalized = " ".join(str(alt_text or "").replace("\n", " ").split())

    if node.tag == f"{{{DOCX_XML_NS['w']}}}drawing":
        for doc_pr in node.findall(".//wp:docPr", DOCX_XML_NS):
            if normalized:
                doc_pr.attrib["descr"] = normalized
            else:
                doc_pr.attrib.pop("descr", None)
            doc_pr.attrib.pop("title", None)

        for nv_props in node.findall(".//pic:cNvPr", DOCX_XML_NS):
            if normalized:
                nv_props.attrib["descr"] = normalized
            else:
                nv_props.attrib.pop("descr", None)
            nv_props.attrib.pop("title", None)
        return

    for shape in node.findall(".//v:shape", DOCX_XML_NS):
        if normalized:
            shape.attrib["alt"] = normalized
            shape.attrib["title"] = normalized
        else:
            shape.attrib.pop("alt", None)
            shape.attrib.pop("title", None)

    for image_data in node.findall(".//v:imagedata", DOCX_XML_NS):
        if normalized:
            image_data.attrib["title"] = normalized
            image_data.attrib[f"{{{DOCX_XML_NS['o']}}}title"] = normalized
        else:
            image_data.attrib.pop("title", None)
            image_data.attrib.pop(f"{{{DOCX_XML_NS['o']}}}title", None)

    for ole_object in node.findall(".//o:OLEObject", DOCX_XML_NS):
        if normalized:
            ole_object.attrib["title"] = normalized
            ole_object.attrib["descr"] = normalized
        else:
            ole_object.attrib.pop("title", None)
            ole_object.attrib.pop("descr", None)


def inject_alt_texts_into_docx(docx_path: Path, workbook_alt_texts: list[str], session_rows: list[dict]) -> tuple[bytes, int]:
    updated_parts: dict[str, bytes] = {}
    applied_count = 0

    with zipfile.ZipFile(docx_path, "r") as source_archive:
        members = source_archive.infolist()
        part_roots, targets = collect_docx_injection_targets(source_archive)
        if len(workbook_alt_texts) != len(session_rows):
            raise ValueError(
                f"The uploaded workbook contains {len(workbook_alt_texts)} ALT row(s), but {len(session_rows)} were expected."
            )

        matched_targets = match_session_rows_to_injection_targets(session_rows, targets)

        for index, row in enumerate(session_rows):
            target = matched_targets[index]
            if target is None:
                continue
            expected_role = str(row.get("role", "") or "").strip().lower()
            if expected_role == "equation" and target["role"] in {"equation", "image"}:
                pass
            elif expected_role and target["role"] != expected_role:
                raise ValueError(
                    f"ALT row {index + 1} does not match the current DOCX structure ({expected_role} expected, found {target['role']})."
                )

            apply_alt_text_to_visual_node(target["node"], workbook_alt_texts[index])
            applied_count += 1

        for part_name, root in part_roots.items():
            updated_parts[part_name] = ET.tostring(root, encoding="utf-8", xml_declaration=True)

        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target_archive:
            for member in members:
                if member.filename in updated_parts:
                    target_archive.writestr(member, updated_parts[member.filename])
                else:
                    target_archive.writestr(member, source_archive.read(member.filename))

    return output.getvalue(), applied_count


