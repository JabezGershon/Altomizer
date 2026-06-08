import io
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image as WorkbookImage
from openpyxl.worksheet.worksheet import Worksheet


SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def supported_excel_filename(filename: str) -> bool:
    return Path(filename or "").suffix.lower() in SUPPORTED_EXCEL_SUFFIXES


def build_excel_merge_output_name(first_filename: str) -> str:
    stem = Path(first_filename or "workbook").stem or "workbook"
    return f"{stem}_merged.xlsx"


def _safe_sheet_title(base_title: str, used_titles: set[str]) -> str:
    cleaned = " ".join(str(base_title or "Sheet").replace("/", " ").replace("\\", " ").split()).strip()
    cleaned = cleaned or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    index = 2
    while candidate.lower() in used_titles:
        suffix = f" {index}"
        candidate = f"{cleaned[: max(1, 31 - len(suffix))]}{suffix}"
        index += 1
    used_titles.add(candidate.lower())
    return candidate


def _copy_cell(source_cell: Cell, target_cell: Cell) -> None:
    if source_cell.data_type == "f":
        target_cell.value = f"={source_cell.value}" if source_cell.value and not str(source_cell.value).startswith("=") else source_cell.value
    else:
        target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)


def _copy_images(source_sheet: Worksheet, target_sheet: Worksheet) -> None:
    for source_image in getattr(source_sheet, "_images", []):
        try:
            image_bytes = source_image._data()
        except Exception:
            continue
        target_image = WorkbookImage(io.BytesIO(image_bytes))
        target_image.width = getattr(source_image, "width", target_image.width)
        target_image.height = getattr(source_image, "height", target_image.height)
        if getattr(source_image, "anchor", None) is not None:
            target_image.anchor = copy(source_image.anchor)
        target_sheet.add_image(target_image)


def _shift_image_anchor(anchor: object, *, row_offset: int, col_offset: int = 0) -> object:
    shifted_anchor = copy(anchor)
    marker_from = getattr(shifted_anchor, "_from", None)
    if marker_from is not None:
        marker_from.row = int(marker_from.row) + int(row_offset)
        marker_from.col = int(marker_from.col) + int(col_offset)
    marker_to = getattr(shifted_anchor, "_to", None)
    if marker_to is not None:
        marker_to.row = int(marker_to.row) + int(row_offset)
        marker_to.col = int(marker_to.col) + int(col_offset)
    return shifted_anchor


def _copy_images_with_row_offset(source_sheet: Worksheet, target_sheet: Worksheet, *, row_offset: int) -> None:
    for source_image in getattr(source_sheet, "_images", []):
        try:
            image_bytes = source_image._data()
        except Exception:
            continue
        target_image = WorkbookImage(io.BytesIO(image_bytes))
        target_image.width = getattr(source_image, "width", target_image.width)
        target_image.height = getattr(source_image, "height", target_image.height)
        anchor = getattr(source_image, "anchor", None)
        if anchor is not None:
            target_image.anchor = _shift_image_anchor(anchor, row_offset=row_offset)
        target_sheet.add_image(target_image)


def _copy_sheet(source_sheet: Worksheet, target_sheet: Worksheet) -> None:
    for row in source_sheet.iter_rows():
        for source_cell in row:
            target_cell = target_sheet.cell(row=source_cell.row, column=source_cell.column)
            _copy_cell(source_cell, target_cell)

    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

    for column_key, column_dimension in source_sheet.column_dimensions.items():
        target_dimension = target_sheet.column_dimensions[column_key]
        target_dimension.width = column_dimension.width
        target_dimension.hidden = column_dimension.hidden
        target_dimension.bestFit = column_dimension.bestFit
        target_dimension.collapsed = column_dimension.collapsed
        target_dimension.outlineLevel = column_dimension.outlineLevel
        target_dimension.min = column_dimension.min
        target_dimension.max = column_dimension.max
        if getattr(column_dimension, "style_id", 0):
            target_dimension._style = copy(column_dimension._style)

    for row_key, row_dimension in source_sheet.row_dimensions.items():
        target_dimension = target_sheet.row_dimensions[row_key]
        target_dimension.height = row_dimension.height
        target_dimension.hidden = row_dimension.hidden
        target_dimension.collapsed = row_dimension.collapsed
        target_dimension.outlineLevel = row_dimension.outlineLevel
        target_dimension.thickTop = row_dimension.thickTop
        target_dimension.thickBot = row_dimension.thickBot
        if getattr(row_dimension, "style_id", 0):
            target_dimension._style = copy(row_dimension._style)

    if source_sheet.freeze_panes:
        target_sheet.freeze_panes = source_sheet.freeze_panes
    if source_sheet.sheet_view:
        target_sheet.sheet_view.zoomScale = source_sheet.sheet_view.zoomScale
        target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines

    target_sheet.sheet_format.defaultColWidth = source_sheet.sheet_format.defaultColWidth
    target_sheet.sheet_format.defaultRowHeight = source_sheet.sheet_format.defaultRowHeight
    target_sheet.sheet_properties.tabColor = copy(source_sheet.sheet_properties.tabColor)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.page_setup = copy(source_sheet.page_setup)
    target_sheet.print_options = copy(source_sheet.print_options)
    target_sheet.views = copy(source_sheet.views)
    _copy_images(source_sheet, target_sheet)


def _copy_concatenated_sheet_rows(
    source_sheet: Worksheet,
    target_sheet: Worksheet,
    *,
    start_row: int,
    row_offset: int,
) -> None:
    for row in source_sheet.iter_rows(min_row=start_row):
        for source_cell in row:
            target_cell = target_sheet.cell(row=source_cell.row + row_offset, column=source_cell.column)
            _copy_cell(source_cell, target_cell)

    for merged_range in source_sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if max_row < start_row or min_row < start_row:
            continue
        target_sheet.merge_cells(
            start_row=min_row + row_offset,
            start_column=min_col,
            end_row=max_row + row_offset,
            end_column=max_col,
        )

    for column_key, column_dimension in source_sheet.column_dimensions.items():
        target_dimension = target_sheet.column_dimensions[column_key]
        if column_dimension.width is not None:
            target_dimension.width = max(float(target_dimension.width or 0), float(column_dimension.width))
        target_dimension.hidden = bool(target_dimension.hidden or column_dimension.hidden)
        target_dimension.bestFit = bool(target_dimension.bestFit or column_dimension.bestFit)
        target_dimension.collapsed = bool(target_dimension.collapsed or column_dimension.collapsed)
        target_dimension.outlineLevel = max(int(target_dimension.outlineLevel or 0), int(column_dimension.outlineLevel or 0))
        if column_dimension.min is not None:
            target_dimension.min = column_dimension.min
        if column_dimension.max is not None:
            target_dimension.max = column_dimension.max
        if getattr(column_dimension, "style_id", 0):
            target_dimension._style = copy(column_dimension._style)

    for row_key, row_dimension in source_sheet.row_dimensions.items():
        if int(row_key) < start_row:
            continue
        target_dimension = target_sheet.row_dimensions[int(row_key) + row_offset]
        target_dimension.height = row_dimension.height
        target_dimension.hidden = row_dimension.hidden
        target_dimension.collapsed = row_dimension.collapsed
        target_dimension.outlineLevel = row_dimension.outlineLevel
        target_dimension.thickTop = row_dimension.thickTop
        target_dimension.thickBot = row_dimension.thickBot
        if getattr(row_dimension, "style_id", 0):
            target_dimension._style = copy(row_dimension._style)

    if target_sheet.freeze_panes is None and source_sheet.freeze_panes:
        target_sheet.freeze_panes = source_sheet.freeze_panes
    if source_sheet.sheet_view:
        target_sheet.sheet_view.zoomScale = source_sheet.sheet_view.zoomScale
        target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines

    if target_sheet.sheet_format.defaultColWidth is None and source_sheet.sheet_format.defaultColWidth is not None:
        target_sheet.sheet_format.defaultColWidth = source_sheet.sheet_format.defaultColWidth
    if target_sheet.sheet_format.defaultRowHeight is None and source_sheet.sheet_format.defaultRowHeight is not None:
        target_sheet.sheet_format.defaultRowHeight = source_sheet.sheet_format.defaultRowHeight
    if target_sheet.sheet_properties.tabColor is None and source_sheet.sheet_properties.tabColor is not None:
        target_sheet.sheet_properties.tabColor = copy(source_sheet.sheet_properties.tabColor)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.page_setup = copy(source_sheet.page_setup)
    target_sheet.print_options = copy(source_sheet.print_options)
    target_sheet.views = copy(source_sheet.views)
    _copy_images_with_row_offset(source_sheet, target_sheet, row_offset=row_offset)


def merge_excel_workbooks(files: list[tuple[str, bytes]]) -> tuple[bytes, dict]:
    if not files:
        raise ValueError("Upload at least one Excel workbook.")

    merged_workbook = Workbook()
    target_sheet = merged_workbook.active
    target_sheet.title = "Merged"
    merged_sheet_count = 0
    merged_file_count = 0
    merged_row_count = 0
    next_target_row = 1

    for filename, workbook_bytes in files:
        if not supported_excel_filename(filename):
            raise ValueError(f"{filename or 'Workbook'} is not a supported Excel file.")
        if not workbook_bytes:
            raise ValueError(f"{filename or 'Workbook'} is empty.")

        try:
            source_workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
        except Exception as exc:
            raise ValueError(f"{filename or 'Workbook'} could not be opened as an Excel workbook: {exc}") from exc

        merged_file_count += 1
        for source_sheet in source_workbook.worksheets:
            merged_sheet_count += 1
            if source_sheet.max_row <= 0:
                continue

            start_row = 1
            if source_sheet.max_row < start_row:
                continue

            row_offset = next_target_row - start_row
            _copy_concatenated_sheet_rows(
                source_sheet,
                target_sheet,
                start_row=start_row,
                row_offset=row_offset,
            )
            copied_rows = max(0, source_sheet.max_row - start_row + 1)
            merged_row_count += copied_rows
            next_target_row += copied_rows

    if merged_row_count == 0 and target_sheet.max_row == 1 and target_sheet["A1"].value is None:
        target_sheet["A1"] = "Merged"

    output = io.BytesIO()
    merged_workbook.save(output)
    summary = {
        "files": merged_file_count,
        "sheets": merged_sheet_count,
        "output_sheets": 1,
        "rows": merged_row_count,
        "sheet_titles": [target_sheet.title],
    }
    return output.getvalue(), summary
